import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import uuid
import json
import pytz
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# Google Calendar API setup
SCOPES = [
    'https://www.googleapis.com/auth/calendar',
    'https://www.googleapis.com/auth/calendar.events'
]

def authenticate_google_calendar():
    creds = None
    if os.path.exists('token.json'):
        with open('token.json', 'r') as token:
            creds = Credentials.from_authorized_user_info(json.load(token), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    service = build('calendar', 'v3', credentials=creds)
    return service

def add_event_to_calendar(task_id, task_name, start_datetime, end_datetime, timezone):
    service = authenticate_google_calendar()
    event = {
        'summary': f'{task_id} - {task_name}',
        'start': {
            'dateTime': start_datetime.isoformat(),
            'timeZone': timezone,
        },
        'end': {
            'dateTime': end_datetime.isoformat(),
            'timeZone': timezone,
        },
    }
    event = service.events().insert(calendarId='primary', body=event).execute()
    return event['id']

def remove_event_from_calendar(event_id):
    service = authenticate_google_calendar()
    service.events().delete(calendarId='primary', eventId=event_id).execute()

# Convert time to 24-hour format
def convert_to_24hour(time, period):
    hour, minute = time.split(":")
    hour = int(hour)
    if period == "PM" and hour != 12:
        hour += 12
    if period == "AM" and hour == 12:
        hour = 0
    return f"{hour:02d}:{minute}"

# Add task information to log
def add_task_to_log(task_name, start_date, start_time, start_period, end_date, end_time, end_period, timezone, task_log):
    try:
        start_time_24 = convert_to_24hour(start_time, start_period)
        start_datetime = datetime.strptime(f"{start_date} {start_time_24}", '%Y-%m-%d %H:%M')

        end_time_24 = convert_to_24hour(end_time, end_period)
        end_datetime = datetime.strptime(f"{end_date} {end_time_24}", '%Y-%m-%d %H:%M')

        # Set timezone
        tz = pytz.timezone(timezone)
        start_datetime = tz.localize(start_datetime)
        end_datetime = tz.localize(end_datetime)

        if end_datetime < start_datetime:
            end_datetime += timedelta(days=1)

        hours_worked = (end_datetime - start_datetime).total_seconds() / 3600

        wb, sheet = open_task_log(task_log)

        task_id = str(uuid.uuid4())
        event_id = add_event_to_calendar(task_id, task_name, start_datetime, end_datetime, timezone)
        task_data = [task_id, task_name, start_date, start_time, start_period, end_date, end_time, end_period, timezone, hours_worked, event_id]

        sheet.append(task_data)
        wb.save(task_log)
        print(f"Task information added to '{task_log}'.")
    except Exception as e:
        print(f"Error adding task to log: {e}")

def open_task_log(task_log):
    headers = ["ID", "Task", "Start Date", "Start Time", "Start AM/PM", "End Date", "End Time", "End AM/PM", "Timezone", "Decimal Hours", "Event ID"]
    if os.path.exists(task_log):
        wb = openpyxl.load_workbook(task_log)
        sheet = wb.active
        # Ensure the correct columns are present
        if sheet.max_row == 0 or sheet[1][0].value != "ID":
            sheet.append(headers)
        else:
            current_headers = [cell.value for cell in sheet[1]]
            if set(headers) != set(current_headers):
                sheet.insert_rows(1)
                sheet.append(headers)
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(headers)
    return wb, sheet

def main():
    task_log = "task_log.xlsx"

    while True:
        task_name = input("Enter task name: ")
        start_date = input("Enter start date (YYYY-MM-DD): ")
        start_time = input("Enter start time (HH:MM): ")
        start_period = input("Enter start time period (AM/PM): ")
        end_date = input("Enter end date (YYYY-MM-DD): ")
        end_time = input("Enter end time (HH:MM): ")
        end_period = input("Enter end time period (AM/PM): ")
        timezone = input("Enter timezone (e.g., America/Detroit): ")

        add_task_to_log(task_name, start_date, start_time, start_period, end_date, end_time, end_period, timezone, task_log)
        print("Task added.\n")

if __name__ == "__main__":
    main()